import streamlit as st
import pandas as pd
from datetime import datetime, date
import os

st.set_page_config(page_title="班级学生管理平台", layout="wide")

# ---------- 密码配置 ----------
TEACHER_PASSWORD = "123456"

# ---------- 数据文件 ----------
DATA_FILE = "class_data.xlsx"

def load_data(sheet_name):
    if os.path.exists(DATA_FILE):
        try:
            df = pd.read_excel(DATA_FILE, sheet_name=sheet_name, engine='openpyxl')
            return df
        except:
            return pd.DataFrame()
    else:
        return pd.DataFrame()

def save_data(df, sheet_name):
    with pd.ExcelWriter(DATA_FILE, engine='openpyxl', mode='a' if os.path.exists(DATA_FILE) else 'w') as writer:
        if os.path.exists(DATA_FILE):
            from openpyxl import load_workbook
            book = load_workbook(DATA_FILE)
            if sheet_name in book.sheetnames:
                idx = book.sheetnames.index(sheet_name)
                book.remove(book.worksheets[idx])
                book.save(DATA_FILE)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def generate_profile(row):
    score = row.get("上学期均分", 0)
    if pd.isna(score):
        score = 0
    tags = str(row.get("特长标签", ""))
    
    if score >= 90:
        base = "🏅 学习顶尖型"
    elif score >= 75:
        base = "📘 良好稳定型"
    elif score >= 60:
        base = "🌱 潜力提升型"
    else:
        base = "⚠️ 需重点关注型"
    
    if "编程" in tags:
        base += " + 逻辑思维强"
    if "足球" in tags or "篮球" in tags:
        base += " + 运动特长"
    if "绘画" in tags or "音乐" in tags:
        base += " + 艺术特长"
    return base

# ---------- 学生端：仅输入姓名 ----------
def student_portal():
    st.header("👨‍🎓 学生个人中心")
    st.info("请输入你的姓名，查看和编辑自己的信息")
    
    student_name = st.text_input("姓名", key="student_login").strip()
    
    if student_name:
        df_students = load_data("students")
        df_tasks = load_data("tasks")
        df_leave = load_data("leaves")
        
        student_data = df_students[df_students["姓名"] == student_name] if not df_students.empty else pd.DataFrame()
        
        if student_data.empty:
            st.warning(f"找不到「{student_name}」同学，请确认姓名是否正确，或联系老师添加")
            return
        
        if len(student_data) > 1:
            st.info(f"找到 {len(student_data)} 位同名同学，请选择你的学号：")
            selected_id = st.selectbox("学号", student_data["学号"].tolist())
            student = student_data[student_data["学号"] == selected_id].iloc[0]
        else:
            student = student_data.iloc[0]
        
        stu_id = student["学号"]
        
        st.success(f"欢迎，{student['姓名']}同学！")
        
        st.subheader("📊 我的画像")
        profile = generate_profile(student)
        st.metric("当前画像", profile)
        
        with st.expander("✏️ 编辑我的信息"):
            col1, col2 = st.columns(2)
            with col1:
                new_tags = st.text_input("特长标签（用逗号分隔）", value=student.get("特长标签", ""))
            with col2:
                new_behavior = st.text_area("最近行为记录", value=student.get("行为记录", ""))
            
            if st.button("更新我的信息"):
                idx = df_students[df_students["学号"] == stu_id].index[0]
                df_students.at[idx, "特长标签"] = new_tags
                df_students.at[idx, "行为记录"] = new_behavior
                df_students.at[idx, "画像描述"] = generate_profile(df_students.iloc[idx])
                save_data(df_students, "students")
                st.success("信息已更新")
                st.rerun()
        
        st.subheader("✅ 我的任务")
        my_tasks = df_tasks[df_tasks["学号"] == stu_id] if not df_tasks.empty else pd.DataFrame()
        if not my_tasks.empty:
            for idx, task in my_tasks.iterrows():
                col1, col2, col3 = st.columns([3, 2, 1])
                with col1:
                    st.write(f"📌 {task['任务名称']}")
                    st.caption(f"截止：{task['截止日期']}")
                with col2:
                    new_status = st.selectbox(
                        "状态", 
                        ["未开始", "进行中", "已完成", "逾期未交"],
                        index=["未开始", "进行中", "已完成", "逾期未交"].index(task["完成状态"]) if task["完成状态"] in ["未开始", "进行中", "已完成", "逾期未交"] else 0,
                        key=f"task_status_{stu_id}_{idx}"
                    )
                with col3:
                    if st.button("更新", key=f"task_update_{stu_id}_{idx}"):
                        df_tasks.at[idx, "完成状态"] = new_status
                        if new_status == "已完成":
                            df_tasks.at[idx, "提交时间"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                        save_data(df_tasks, "tasks")
                        st.success("状态已更新")
                        st.rerun()
        else:
            st.info("暂无任务安排")
        
        st.subheader("📋 请假申请")
        with st.form("student_leave"):
            col1, col2 = st.columns(2)
            with col1:
                leave_date = st.date_input("请假日期", value=date.today())
                periods = st.multiselect("请假节次", ["第1节", "第2节", "第3节", "第4节", "第5节", "第6节"])
            with col2:
                reason = st.text_area("事由")
            
            submitted = st.form_submit_button("提交请假申请")
            if submitted and periods:
                df_leave = load_data("leaves")
                new_row = pd.DataFrame([{
                    "学号": stu_id,
                    "姓名": student["姓名"],
                    "请假日期": leave_date.strftime("%Y-%m-%d"),
                    "节次": ",".join(periods),
                    "事由": reason,
                    "预审状态": "预审中",
                    "班主任意见": ""
                }])
                df_leave = pd.concat([df_leave, new_row], ignore_index=True)
                save_data(df_leave, "leaves")
                st.success("请假已提交，等待审批")
                st.rerun()
        
        with st.expander("📋 我的请假记录"):
            my_leaves = df_leave[df_leave["学号"] == stu_id] if not df_leave.empty else pd.DataFrame()
            if not my_leaves.empty:
                st.dataframe(my_leaves[["请假日期", "节次", "事由", "预审状态", "班主任意见"]], use_container_width=True)
            else:
                st.info("暂无请假记录")
        
        with st.expander("🏆 比赛报名"):
            with st.form("student_comp"):
                comp_name = st.text_input("比赛名称")
                if st.form_submit_button("提交报名申请"):
                    if comp_name:
                        df_comp = load_data("competitions")
                        new_row = pd.DataFrame([{
                            "学号": stu_id,
                            "姓名": student["姓名"],
                            "比赛名称": comp_name,
                            "报名时间": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "审核状态": "待审核"
                        }])
                        df_comp = pd.concat([df_comp, new_row], ignore_index=True)
                        save_data(df_comp, "competitions")
                        st.success("报名已提交，等待审核")
                        st.rerun()
                    else:
                        st.error("请输入比赛名称")

# ---------- 教师端 ----------
def teacher_login():
    st.header("🔐 教师后台登录")
    password = st.text_input("请输入管理员密码", type="password")
    if password == TEACHER_PASSWORD:
        st.session_state["teacher_logged_in"] = True
        st.rerun()
    elif password:
        st.error("密码错误")

def teacher_portal():
    st.header("📊 教师后台管理")
    
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "学生信息", "比赛报名", "请假审批", "任务管理", "数据导出", "综合看板"
    ])
    
    with tab1:
        st.subheader("👥 学生基本管理")
        df_students = load_data("students")
        
        with st.expander("➕ 添加新学生"):
            col1, col2, col3 = st.columns(3)
            with col1:
                new_id = st.text_input("学号")
                new_name = st.text_input("姓名")
            with col2:
                new_gender = st.selectbox("性别", ["男", "女"])
                new_group = st.text_input("小组")
            with col3:
                new_score = st.number_input("上学期均分", 0, 100, 80)
            
            if st.button("添加学生"):
                if new_id and new_name:
                    new_row = pd.DataFrame([{
                        "学号": new_id, "姓名": new_name, "性别": new_gender,
                        "小组": new_group, "上学期均分": new_score,
                        "特长标签": "", "行为记录": "", "画像描述": ""
                    }])
                    df_students = pd.concat([df_students, new_row], ignore_index=True)
                    df_students["画像描述"] = df_students.apply(generate_profile, axis=1)
                    save_data(df_students, "students")
                    st.success(f"已添加 {new_name}")
                    st.rerun()
        
        edited = st.data_editor(df_students, num_rows="dynamic", use_container_width=True)
        if st.button("保存学生信息修改"):
            edited["画像描述"] = edited.apply(generate_profile, axis=1)
            save_data(edited, "students")
            st.success("已保存")
    
    with tab2:
        st.subheader("🏆 比赛报名管理")
        df_comp = load_data("competitions")
        if df_comp.empty:
            df_comp = pd.DataFrame(columns=["学号", "姓名", "比赛名称", "报名时间", "审核状态"])
        
        edited_comp = st.data_editor(df_comp, num_rows="dynamic", use_container_width=True)
        if st.button("保存比赛报名"):
            save_data(edited_comp, "competitions")
            st.success("已保存")
    
    with tab3:
        st.subheader("📋 请假审批")
        df_leave = load_data("leaves")
        
        if not df_leave.empty:
            pending = df_leave[df_leave["预审状态"] == "预审中"]
            if not pending.empty:
                for idx, row in pending.iterrows():
                    with st.expander(f"{row['姓名']} - {row['请假日期']} 请假{row['节次']}"):
                        st.write(f"事由：{row['事由']}")
                        new_status = st.selectbox("审批意见", ["已批准", "已拒绝"], key=f"leave_{idx}")
                        comment = st.text_input("班主任意见", key=f"comment_{idx}")
                        if st.button("确认审批", key=f"btn_{idx}"):
                            df_leave.at[idx, "预审状态"] = new_status
                            df_leave.at[idx, "班主任意见"] = comment
                            save_data(df_leave, "leaves")
                            st.success("已处理")
                            st.rerun()
            else:
                st.info("暂无待审批请假")
            
            with st.expander("全部请假记录"):
                st.dataframe(df_leave, use_container_width=True)
        else:
            st.info("暂无请假记录")
    
    with tab4:
        st.subheader("✅ 任务管理")
        df_tasks = load_data("tasks")
        
        with st.form("new_task"):
            col1, col2 = st.columns(2)
            with col1:
                stu_id = st.text_input("学号", key="task_stu")
                stu_name = st.text_input("姓名", key="task_name")
            with col2:
                task_title = st.text_input("任务名称")
                due_date = st.date_input("截止日期")
            
            if st.form_submit_button("添加任务"):
                if stu_id and task_title:
                    new_row = pd.DataFrame([{
                        "学号": stu_id, "姓名": stu_name, "任务名称": task_title,
                        "截止日期": due_date.strftime("%Y-%m-%d"), "完成状态": "未开始",
                        "提交时间": "", "备注": ""
                    }])
                    df_tasks = pd.concat([df_tasks, new_row], ignore_index=True)
                    save_data(df_tasks, "tasks")
                    st.success("已添加")
                    st.rerun()
        
        edited_tasks = st.data_editor(df_tasks, num_rows="dynamic", use_container_width=True)
        if st.button("保存任务"):
            save_data(edited_tasks, "tasks")
            st.success("已保存")
    
    with tab5:
        st.subheader("📎 数据导出与下载")
        df_students = load_data("students")
        df_comp = load_data("competitions")
        df_leave = load_data("leaves")
        df_tasks = load_data("tasks")
        
        if not df_students.empty:
            csv = df_students.to_csv(index=False).encode('utf-8-sig')
            st.download_button("📥 导出学生信息", csv, "学生信息.csv", "text/csv")
        if not df_comp.empty:
            csv = df_comp.to_csv(index=False).encode('utf-8-sig')
            st.download_button("📥 导出比赛报名", csv, "比赛报名.csv", "text/csv")
        if not df_leave.empty:
            csv = df_leave.to_csv(index=False).encode('utf-8-sig')
            st.download_button("📥 导出请假记录", csv, "请假记录.csv", "text/csv")
        if not df_tasks.empty:
            csv = df_tasks.to_csv(index=False).encode('utf-8-sig')
            st.download_button("📥 导出任务数据", csv, "任务数据.csv", "text/csv")
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "rb") as f:
                st.download_button("💾 备份完整数据", f, f"班级数据备份_{datetime.now().strftime('%Y%m%d')}.xlsx")
    
    with tab6:
        st.subheader("📈 班级综合看板")
        df_students = load_data("students")
        df_comp = load_data("competitions")
        df_tasks = load_data("tasks")
        
        if not df_students.empty:
            st.subheader("学生画像分布")
            portrait_counts = df_students["画像描述"].value_counts()
            st.bar_chart(portrait_counts)
            
            if not df_tasks.empty:
                st.subheader("任务完成情况")
                task_status = df_tasks["完成状态"].value_counts()
                st.bar_chart(task_status)
            
            col1, col2, col3 = st.columns(3)
            col1.metric("总学生数", len(df_students))
            col2.metric("平均成绩", f"{df_students['上学期均分'].mean():.1f}")
            col3.metric("比赛报名数", len(df_comp) if not df_comp.empty else 0)

# ---------- 主入口 ----------
def main():
    st.sidebar.title("导航")
    
    if "teacher_logged_in" not in st.session_state:
        st.session_state["teacher_logged_in"] = False
    
    role = st.sidebar.radio("登录身份", ["学生入口", "教师后台"])
    
    if role == "学生入口":
        student_portal()
    else:
        if st.session_state["teacher_logged_in"]:
            teacher_portal()
            if st.sidebar.button("退出登录"):
                st.session_state["teacher_logged_in"] = False
                st.rerun()
        else:
            teacher_login()

if __name__ == "__main__":
    main()